import openpyxl
import requests
import json

url = "http://localhost:3000/api/v1/course"

httpheaders = {
    "User-Agent": "curl/7.81.0",
    "Accept": "*/*",
    "Content-Type" : "application/json"
}

file_path = "../data/CSE_Courses.xlsx"
wb = openpyxl.load_workbook(file_path, data_only=True)

sheet_name="Sheet1"
ws = wb[sheet_name] if sheet_name else wb.active

headers = []
for cell in next(ws.iter_rows(min_row=1, max_row=1)):
    if cell.value is not None:
        headers.append(cell.value.strip().lower().replace(" ", "_"))
    else:
        headers.append("")  # Or use a placeholder like f"column_{i}"

expected_keys = {"course_id", "course_name", "course_category", "offered_department_id","course_coordinator_id","course_champion_id","caia_coordinator_id" }
    
# Validate headers
if not expected_keys.issubset(set(headers)):
  raise ValueError(f"Expected columns {expected_keys} not found in sheet headers: {headers}")

# Read data rows and convert to dictionary
data = []
for row in ws.iter_rows(min_row=2, values_only=True):
  course_dict = dict(zip(headers, row))
  #data.append(emp_dict)
  print(course_dict)

  
  print(json.dumps(course_dict))
  response = requests.post(url, data=json.dumps(course_dict), headers=httpheaders, verify=False)

  # Printing the response
  print("Status Code:", response.status_code)
  print("Response Body:", response.text)  

"""
# Data to be sent in the POST request
payload = {
    "course_id"             : "CSE3122",
    "course_name"           : "Operating Systems",
    "course_category"       : "PC",
    "offered_department_id" : "301",
    "course_coordinator_id" : "700917",
    "course_champion_id"    : "700917",
    "caia_coordinator_id"   : "700917"
}
"""