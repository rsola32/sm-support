import openpyxl
import requests
import json

# URL to send the POST request to
url = "http://localhost:3000/api/v1/faculty"

httpheaders = {
    "User-Agent": "curl/7.81.0",
    "Accept": "*/*",
    "Content-Type" : "application/json"
}


file_path = "../data/CSE_Faculty_List.xlsx"
wb = openpyxl.load_workbook(file_path, data_only=True)

sheet_name="Sheet1"
ws = wb[sheet_name] if sheet_name else wb.active

headers = []
for cell in next(ws.iter_rows(min_row=1, max_row=1)):
    if cell.value is not None:
        headers.append(cell.value.strip().lower().replace(" ", "_"))
    else:
        headers.append("")  # Or use a placeholder like f"column_{i}"

expected_keys = {"employee_id", "employee_name", "designation", "department_id","employee_mail_id"}
    
# Validate headers
if not expected_keys.issubset(set(headers)):
  raise ValueError(f"Expected columns {expected_keys} not found in sheet headers: {headers}")

# Read data rows and convert to dictionary
data = []
for row in ws.iter_rows(min_row=2, values_only=True):
  emp_dict = dict(zip(headers, row))
  #data.append(emp_dict)
  if 'doj' not in emp_dict or emp_dict['doj']==None:
     emp_dict['doj']="2025-01-01"
  emp_dict['password']="Gitam123"
  print(emp_dict)

  
  print(json.dumps(emp_dict))
  # Sending the POST request
  response = requests.post(url, data=json.dumps(emp_dict), headers=httpheaders, verify=False)

  # Printing the response
  print("Status Code:", response.status_code)
  print("Response Body:", response.text)

"""
# Data to be sent in the POST request
payload = {
  "employee_id": "700918",
  "employee_email": "ramx@gitam.edu",
  "employee_name": "Ram X",
  "designation": "Assistant Professor",
  "joining_date": "2024-07-31",
  "department_id": "301",
  "password": "caia123"
}
"""