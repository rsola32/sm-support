import openpyxl
import requests
import json

# URL to send the POST request to
skill_url = "http://localhost:3000/api/v1/skill"
subskill_url = "http://localhost:3000/api/v1/subskill"
mapping_url = "http://localhost:3000/api/v1/course_skill_mapping"


httpheaders = {
    "User-Agent": "curl/7.81.0",
    "Accept": "*/*",
    "Content-Type" : "application/json"
}

file_path = "../data/EECE_Consolidated_Skill_Repo.xlsx"
wb = openpyxl.load_workbook(file_path, data_only=True)

sheet_name="Sheet1"
ws = wb[sheet_name] if sheet_name else wb.active

course_id = ws.cell(2,1).value
print("ID:",course_id)

skill_data = []
skill_set = set()
seq = 0
sseq = 0

for row in ws.iter_rows(min_row=2, values_only=True):
  skill_type = row[1]   # Column B
  skill = row[2]        # Column C
  sub_skill = row[3]    # Column D
  description = row[4]  # Column E
  is_required = row[5]  # Column E

  print(row)
  #flag = False
  #if skill not in skill_set:
  skill_payload = {
    "skill_id" : f"{course_id}S{seq}",
    "skill_name" : skill,
    "skill_description " : "TODO"
  }
  seq = seq + 1
  skill_set.add(skill)
     #print(json.dumps(skill_payload))
     #response = requests.post(skill_url,data=json.dumps(skill_payload), headers=httpheaders, verify=False)
     #print("Status Code:", response.status_code)
     #print("Response Body:", response.text)
     #print(skill_payload)
     #flag = True   
        
  subskill_payload = {
    "subskill_id" : f"{skill_payload["skill_id"]}x{sseq}",
    "subskill_name" : sub_skill,
    "subskill_description" : description,
    "skill_id" : skill_payload["skill_id"]
  }
  subskill_payload['required'] = is_required=="Must"
  sseq = sseq + 1
  response = requests.post(subskill_url, data=json.dumps(subskill_payload), headers=httpheaders, verify=False)
  print("Status Code:", response.status_code, sub_skill)
  print("Response Body:", response.text)
  print(subskill_payload)

  mapping_payload = {
    "skill_id" : skill_payload["skill_id"],
    "course_id" : course_id,
    "skill_type" : skill_type
  }
  print(mapping_payload)
  #if flag:
  #response = requests.post(mapping_url, data=json.dumps(mapping_payload), headers=httpheaders, verify=False)
  #print("Status Code:", response.status_code, skill, course_id)
  #print("Response Body:", response.text)
 
# for entry in skill_data:
#  print(entry)

  
  # Sending the POST request
  #response = requests.post(url, data=json.dumps(emp_dict), headers=httpheaders, verify=False)

  # Printing the response
  #print("Status Code:", response.status_code)
  #print("Response Body:", response.text)

"""
# Data to be sent in the POST request
payload = {
  "employee_id": "700918",
  "employee_email": "ramx@gitam.edu",
  "employee_name": "Ram X",
  "designation": "Assistant Professor",
  "joining_date": "2024-07-31",
  "department_id": "301",
  "password": "caia123"
}
"""
